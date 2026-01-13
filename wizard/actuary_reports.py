import xlwt
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
import base64
import time
import datetime
import io
from io import StringIO
from datetime import datetime
from datetime import timedelta
from odoo.exceptions import UserError
from odoo import api, fields, models,_

class actuary_reports(models.TransientModel):
    _name= "actuary.reports"
    _description= "Actuary reports wizard"

    action = fields.Selection([
        ('enrollee', 'Enrollee Membership'),
        ('claims_paid', 'Paid Claims'),
        ('claims_out', 'Outstanding Claims'),
        ('claims_inc', 'Incurred Claims'),
        ('line', 'Claims Procedures'),
        ('pay', 'Procedure Payments'),
        ('diag', 'Diagnosis'),
        ('cap', 'Capitation Data'),
        ('prem', 'Premium Data'),
        ('refund','Refunds'),
        ], string='Select Report', copy=False)
    status = fields.Selection([
        ('all', 'All'),
        ('active', 'Active'),
        ('inactive', 'Inactive'),
        ], string='Status', copy=False)
    start_d = fields.Date(string='Start Date', required=True)
    end_d = fields.Date(string='End Date')
    
    def replacer(self,inputstr):
        s1 = inputstr.replace('en_US','')
        s2 = s1.replace(':','')
        s3 = s2.replace('}','')
        s4 = s3.replace('{','')
        s5 = s4.replace("'","")
        return s5.strip()
    
    def actuary_report(self):
        stat=" "
        endd=""
        if self.action == 'enrollee':
            if self.end_d:
                endd = " and e.start_date <= '" + str(self.end_d) + "'"
        elif self.action == 'pay':
            if self.end_d:
                endd = " and i.write_date <= '" + str(self.end_d) + "'"
        elif self.action in ('line','claims_paid','diag','claims_out','claims_inc'):
            if self.end_d:
                endd = " and o.date_order <= '" + str(self.end_d) + "'"
        elif self.action == 'cap':
            if self.end_d:
                endd = "  (o.start_date between '" + str(self.start_d) + "' and '" + str(self.end_d) + "' or o.end_date between '" + str(self.start_d) + "' and '" + str(self.end_d) + "' or ('" + str(self.start_d) + "' between o.start_date and o.end_date and '" + str(self.end_d) + "' between o.start_date and o.end_date))"
            else:
                endd = "  (o.start_date >= '" + str(self.start_d) + "' or o.end_date >= '" + str(self.start_d) + "' or '" + str(self.start_d) + "' between o.start_date and o.end_date)"
        elif self.action == 'prem':
            if self.end_d:
                endd = "  (so.start_date between '" + str(self.start_d) + "' and '" + str(self.end_d) + "' or so.end_date between '" + str(self.start_d) + "' and '" + str(self.end_d) + "' or ('" + str(self.start_d) + "' between so.start_date and so.end_date and '" + str(self.end_d) + "' between so.start_date and so.end_date))"
            else:
                endd = "  (so.start_date >= '" + str(self.start_d) + "' or so.end_date >= '" + str(self.start_d) + "' or '" + str(self.start_d) + "' between so.start_date and so.end_date)"
        if self.status=='active':
            stat=" and e.active = 'TRUE' " 
        elif self.status=='inactive':
            stat=" and e.active = 'FALSE' "
        sql_enrollee="select distinct e.surname, e.firstname, e.othername,e.code, c.name employer, c.id cicode, o.name as policy, e.dob, e.gender, e.marital,e.start_date, to_char(e.start_date,'MM-YYYY') mthyr, e.end_date, o.end_date pend_date, e.active, t.name, l.price_unit, l.product_uom_qty duration, e.coverage, e.type, h.name hcp,h.id hcpid, e.town, h.city, (select sum(dep1 +  dep2 +  dep3 + dep4 + dep5 + ext1 + ext2 + ext3 + ext4 + ext5 ) from (select sum(CASE WHEN w.dependent1 is not null and (select f.active from enrollee f where id = w.dependent1) = True and (select g.dob from enrollee g where g.id = w.dependent1) <= Now() - interval '18 years' THEN 1 ELSE 0 END) dep1, sum(CASE WHEN w.dependent2  is not null and (select f.active from enrollee f where id = w.dependent2) = True and (select g.dob from enrollee g where g.id = w.dependent2) <= Now() - interval '18 years' THEN 1 else 0 END)  dep2, sum(CASE WHEN w.dependent3  is not null and (select f.active from enrollee f where id = w.dependent3) = True  and (select g.dob from enrollee g where g.id = w.dependent3) <= Now() - interval '18 years' THEN 1 else 0 END) dep3, sum(CASE WHEN w.dependent4  is not null and (select f.active from enrollee f where id = w.dependent4) = True and (select g.dob from enrollee g where g.id = w.dependent4) <= Now() - interval '18 years' THEN 1 else 0 END)  dep4, sum(CASE WHEN w.dependent5  is not null and (select f.active from enrollee f where id = w.dependent5) = True  and (select g.dob from enrollee g where g.id = w.dependent5) <= Now() - interval '18 years' THEN 1 else 0 END) dep5, sum(CASE WHEN w.extra1  is not null and (select f.active from enrollee f where id = w.extra1) = True THEN 1 else 0 END) ext1, sum(CASE WHEN w.extra2  is not null and (select f.active from enrollee f where id = w.extra2) = True THEN 1 else 0 END) ext2, sum(CASE WHEN w.extra3  is not null and (select f.active from enrollee f where id = w.extra3) = True THEN 1 else 0 END) ext3, sum(CASE WHEN w.extra4  is not null and (select f.active from enrollee f where id = w.extra4) = True THEN 1 else 0 END) ext4, sum(CASE WHEN w.extra5  is not null and (select f.active from enrollee f where id = w.extra5) = True THEN 1 else 0 END) ext5 from enrollee w where w.id = e.id) as cnt) as adult_dep_cnt, (select sum(dep1 +  dep2 +  dep3 + dep4 + dep5 + ext1 + ext2 + ext3 + ext4 + ext5 ) from (select  sum(CASE WHEN w.dependent1 is not null and (select f.active from enrollee f where id = w.dependent1) = True and (select g.dob from enrollee g where g.id = w.dependent1) >= Now() - interval '18 years' THEN 1 ELSE 0 END) dep1, sum(CASE WHEN w.dependent2  is not null and (select f.active from enrollee f where id = w.dependent2) = True  and (select g.dob from enrollee g where g.id = w.dependent2) >= Now() - interval '18 years' THEN 1 else 0 END)  dep2, sum(CASE WHEN w.dependent3  is not null and (select f.active from enrollee f where id = w.dependent3) = True  and (select g.dob from enrollee g where g.id = w.dependent3) >= Now() - interval '18 years' THEN 1 else 0 END) dep3,sum(CASE WHEN w.dependent4  is not null and (select f.active from enrollee f where id = w.dependent4) = True  and (select g.dob from enrollee g where g.id = w.dependent4) >= Now() - interval '18 years' THEN 1 else 0 END)  dep4, sum(CASE WHEN w.dependent5  is not null and (select f.active from enrollee f where id = w.dependent5) = True and (select g.dob from enrollee g where g.id = w.dependent5) >= Now() - interval '18 years' THEN 1 else 0 END) dep5, sum(CASE WHEN w.extra1  is not null and (select f.active from enrollee f where id = w.extra1) = True THEN 1 else 0 END) ext1, sum(CASE WHEN w.extra2  is not null and (select f.active from enrollee f where id = w.extra2) = True THEN 1 else 0 END) ext2, sum(CASE WHEN w.extra3  is not null and (select f.active from enrollee f where id = w.extra3) = True THEN 1 else 0 END) ext3, sum(CASE WHEN w.extra4  is not null and (select f.active from enrollee f where id = w.extra4) = True THEN 1 else 0 END) ext4, sum(CASE WHEN w.extra5  is not null and (select f.active from enrollee f where id = w.extra5) = True THEN 1 else 0 END) ext5 from enrollee w where w.id = e.id) as cnt) as child_deps_cnt, case when e.uncapitated = True then 'FFS' else 'Capitation' end as cap_ffs from enrollee e inner join res_partner c on c.id = e.employer left join sale_order_line l on l.enrollee_id = e.id left join sale_order o on o.id = l.order_id inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id  left join res_partner h on h.id = e.hcp where employer <> 7875 and e.start_date >= '" + str(self.start_d) + "' " + endd + stat
        sql_claims_paid="select distinct e.code, e.surname, e.firstname, e.othername, c.id id, o.date_order, c.receipt_date,  o.partner_id hcp, o.name ccode, o.admission_date, o.discharge_date, t.name, o.provider_total, o.amount_total, o.write_date, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7875 then 'NHIS' else 'Private' end categ, case when o.admission_date is null then 'Out' else 'In' end inoutw, (select date_part('day', age(o.discharge_date, o.admission_date))) duration, case when o.partner_id = e.hcp then 'Treatment' else 'Referral' end referal, case when o.partner_id = 8094 then 'Chronic' else 'Non-Chronic' end chronic, c.name reg,i.account_invoice_id, k.payment_date, (select start_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as start_date, (select so.name from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as policy,(select end_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by end_date desc limit 1) as end_date, (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis from purchase_order o inner join enrollee e on o.enrollee_id = e.id left join claims_registration c on c.id = claim_reg_id  inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp left join account_invoice_purchase_order_rel i on i.purchase_order_id = o.id left join account_invoice_payment_rel r on r.invoice_id = i.account_invoice_id left join account_payment k on k.id = r.payment_id left join account_invoice inv on inv.id = r.invoice_id where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and inv.state='paid' and o.date_order >= '" + str(self.start_d) + "' " + endd + " union select e.code, e.surname, e.firstname, e.othername, c.id id, o.date_order, c.receipt_date,  o.partner_id hcp, o.name ccode, o.admission_date, o.discharge_date, t.name, o.provider_total, o.amount_total, o.write_date, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7875 then 'NHIS' else 'Private' end categ, case when o.admission_date is null then 'Out' else 'In' end inoutw, (select date_part('day', age(o.discharge_date, o.admission_date))) duration, case when o.partner_id = e.hcp then 'Treatment' else 'Referral' end referal, case when o.partner_id = 8094 then 'Chronic' else 'Non-Chronic' end chronic, c.name reg, o.id,o.date_order dorder, (select start_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as start_date, (select so.name from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as policy,(select end_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by end_date desc limit 1) as end_date, (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis from purchase_order o inner join enrollee e on o.enrollee_id = e.id left join claims_registration c on c.id = claim_reg_id  inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and o.partner_id = 8094 and o.date_order >= '" + str(self.start_d) + "' " + endd + " order by date_order"
        sql_claims_outstanding="select distinct e.code, e.surname, e.firstname, e.othername, o.date_order, c.receipt_date,o.partner_id hcp, o.name ccode, t.name, o.amount_total, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7065 then 'NHIS' else 'Private' end categ, (select date_part('day', age(o.discharge_date, o.admission_date))) duration, case when o.partner_id = e.hcp then 'Treatment' else 'Referral' end referal, i.account_invoice_id,(select start_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as start_date, (select so.name from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as policy,(select end_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by end_date desc limit 1) as end_date, (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis from purchase_order o inner join enrollee e on o.enrollee_id = e.id  left join claims_registration c on c.id = claim_reg_id inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp left join account_invoice_purchase_order_rel i on i.purchase_order_id = o.id left join account_invoice inv on inv.id = i.account_invoice_id where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and inv.state in ('open','draft') and o.date_order >= '" + str(self.start_d) + "' " + endd + " order by o.date_order"
        sql_pay ="select l.name drug,o.date_order,e.code enrollee, e.surname, e.firstname, e.othername, o.name claim, i.state, h.name hcp, c.receipt_date, g.name employer, case when i.state = 'paid' then i.write_date else NULL end as payment_date ,l.price_subtotal from account_invoice_line l inner join purchase_order_line pl on pl.id=l.purchase_line_id inner join purchase_order o on o.id = pl.order_id  left join account_invoice i on l.invoice_id = i.id inner join enrollee e on e.id = o.enrollee_id left join res_partner h on h.id = e.hcp inner join res_partner g on g.id = e.employer left join claims_registration c on c.id = o.claim_reg_id where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and i.write_date >= '" + str(self.start_d) + "' " + endd + " order by i.write_date"
        sql_lines ="select l.name drug,o.date_order,e.code enrollee,e.surname, e.firstname, e.othername, o.name claim, g.name employer, l.price_total from purchase_order_line l inner join purchase_order o on o.id = l.order_id inner join enrollee e on e.id = o.enrollee_id left join res_partner h on h.id = e.hcp inner join res_partner g on g.id = e.employer left join claims_registration c on c.id = o.claim_reg_id where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and o.date_order >= '" + str(self.start_d) + "' " + endd + " order by o.date_order"
        sql_diag="select diagnosis, details, x.name,o.date_order,e.code enrollee,e.surname, e.firstname, e.othername,o.name claim from claims_diagnosis x inner join purchase_order o on o.id = x.claim_id inner join enrollee e on e.id = o.enrollee_id  where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and o.date_order >= '" + str(self.start_d) + "' " + endd + " and (diagnosis is not null or details is not null) order by o.date_order"
        sql_cap = "select e.code,surname,firstname, othername,emp.name employer, emp.id, t.name as plan, h.name hcp, h.id hid, case when h.cap_amount > 0 then h.cap_amount else t.cap_amount end cap_amount, o.name as policy from enrollee e inner join res_partner h on h.id = e.hcp inner join res_partner emp on emp.id = e.employer left join sale_order_line l on l.enrollee_id = e.id left join sale_order o on o.id = l.order_id inner join product_product p on p.id = l.product_id inner join product_template t on p.product_tmpl_id = t.id where" + endd + " and e.active=True and (e.uncapitated=False or e.uncapitated is null) order by e.code"
        sql_prem = "select e.code, e.surname, e.firstname, e.othername, t.name, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7065 then 'NHIS' else 'Private' end categ, sol.price_subtotal, so.start_date, so.name as policy, so.end_date from sale_order_line sol inner join enrollee e on sol.enrollee_id = e.id inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp inner join sale_order so on sol.order_id = so.id where so.state in ('draft','sale')  and so.doc_type = 'policy' and " + endd + "order by so.start_date"
        sql_claims_incurred_2="select distinct e.code, e.surname, e.firstname, e.othername, o.date_order, c.receipt_date,o.partner_id hcp, o.name ccode, t.name, o.amount_total, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7065 then 'NHIS' else 'Private' end categ, (select date_part('day', age(o.discharge_date, o.admission_date))) duration, case when o.partner_id = e.hcp then 'Treatment' else 'Referral' end referal, (select start_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as start_date, (select so.name from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as policy,(select end_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by end_date desc limit 1) as end_date, (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis from purchase_order o inner join enrollee e on o.enrollee_id = e.id  left join claims_registration c on c.id = claim_reg_id inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp where o.state in ('purchase','done','to approve')  and o.claim_type in ('claim','refund') and o.date_order >= '" + str(self.start_d) + "' " + endd + " order by o.date_order"
        sql_refunds_incurred="select distinct e.code, e.surname, e.firstname, e.othername, o.date_order, c.receipt_date,o.partner_id hcp, o.name ccode, t.name, o.amount_total, z.name employer, h.name provider, e.dob,e.gender, e.occupation, e.type, case when z.id=7065 then 'NHIS' else 'Private' end categ, (select date_part('day', age(o.discharge_date, o.admission_date))) duration, case when o.partner_id = e.hcp then 'Treatment' else 'Referral' end referal, (select start_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as start_date, (select so.name from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by start_date desc limit 1) as policy,(select end_date from sale_order so inner join sale_order_line ol on ol.order_id = so.id where e.id in (select enrollee_id from sale_order_line sol inner join sale_order soinner on sol.order_id = soinner.id where o.date_order between soinner.start_date and soinner.end_date) order by end_date desc limit 1) as end_date, (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis from purchase_order o inner join enrollee e on o.enrollee_id = e.id  left join claims_registration c on c.id = claim_reg_id inner join product_product p on p.id = e.plan inner join product_template t on t.id = p.product_tmpl_id inner join res_partner z on e.employer = z.id inner join res_partner h on h.id = e.hcp where o.state in ('purchase','done','to approve')  and o.claim_type in ('refund') and o.date_order >= '" + str(self.start_d) + "' " + endd + " order by o.date_order"
        sql_claims_incurred="""SELECT DISTINCT
            e.code,
            e.surname,
            e.firstname,
            e.othername,
            o.date_order,
            c.receipt_date,
            o.partner_id AS hcp,
            o.name AS ccode,
            t.name,
            o.amount_total,
            z.name AS employer,
            h.name AS provider,
            e.dob,
            e.gender,
            e.occupation,
            e.type,
            CASE WHEN z.id = 7065 THEN 'NHIS' ELSE 'Private' END AS categ,
            EXTRACT(DAY FROM AGE(o.discharge_date, o.admission_date)) AS duration,
            CASE WHEN o.partner_id = e.hcp THEN 'Treatment' ELSE 'Referral' END AS referal,
            subquery.start_date,
            subquery.policy,
            subquery.end_date,
            (select COALESCE(NULLIF(diagnosis,''), details) from claims_diagnosis where claim_id = o.id limit 1) diagnosis
        FROM purchase_order o
        INNER JOIN enrollee e ON o.enrollee_id = e.id
        LEFT JOIN claims_registration c ON c.id = o.claim_reg_id
        INNER JOIN product_product p ON p.id = e.plan
        INNER JOIN product_template t ON t.id = p.product_tmpl_id
        INNER JOIN res_partner z ON e.employer = z.id
        INNER JOIN res_partner h ON h.id = e.hcp
        LEFT JOIN claims_diagnosis cd on cd.claim_id = o.id
        LEFT JOIN (
            select so.name as policy,so.end_date,so.start_date,ol.enrollee_id from sale_order so 
            inner join sale_order_line ol on ol.order_id = so.id 
            inner join enrollee en on en.id = ol.enrollee_id limit 1
        ) AS subquery ON e.id = subquery.enrollee_id and o.date_order between subquery.start_date and subquery.end_date
        WHERE o.state IN ('purchase', 'done', 'to approve')
            AND o.claim_type IN ('claim', 'refund')
            AND o.date_order >= '""" + str(self.start_d) + """' """ + endd + """ ORDER BY o.date_order;
        """
        if self.action == 'enrollee':
            self.env.cr.execute(sql_enrollee)
            res = self.env.cr.dictfetchall()
            return self._print_report_excel(res)
        elif self.action == 'claims_paid':
            self.env.cr.execute(sql_claims_paid)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel2(res)
        elif self.action == 'pay':
            self.env.cr.execute(sql_pay)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel3(res)
        elif self.action == 'line':
            self.env.cr.execute(sql_lines)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel4(res)
        elif self.action == 'cap':
            self.env.cr.execute(sql_cap)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel_cap(res)
        elif self.action == 'claims_out':
            self.env.cr.execute(sql_claims_outstanding)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel_claim_out(res)
        elif self.action == 'claims_inc':
            self.env.cr.execute(sql_claims_incurred)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel_claim_incurred_multi(res)
        elif self.action == 'refund':
            self.env.cr.execute(sql_refunds_incurred)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel_claim_incurred(res)
        elif self.action == 'prem':
            self.env.cr.execute(sql_prem)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel_prem(res)
        else:
            self.env.cr.execute(sql_diag)
            res = self.env.cr.dictfetchall()	
            return self._print_report_excel5(res)

    def _print_report_excel(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Enrollee Membership'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)     
        
        sheet.write(8, 0, 'Surname',title_style1_table_head)
        sheet.write(8, 1, 'Firstname',title_style1_table_head)
        sheet.write(8, 2, 'Othername',title_style1_table_head)
        sheet.write(8, 3, 'Enrollee Code',title_style1_table_head)
        sheet.write(8, 4, 'Employer',title_style1_table_head)
        sheet.write(8, 5, 'Employer ID',title_style1_table_head)
        sheet.write(8, 6, 'Policy No',title_style1_table_head)
        sheet.write(8, 7, 'DoB',title_style1_table_head)
        sheet.write(8, 8, 'Gender',title_style1_table_head)
        sheet.write(8, 9, 'Marital Status',title_style1_table_head)
        sheet.write(8, 10, 'Date of Registration',title_style1_table_head)
        sheet.write(8, 11, 'Mth-Year of Registration',title_style1_table_head)
        sheet.write(8, 12, 'Termination Date',title_style1_table_head)
        sheet.write(8, 13, 'Active',title_style1_table_head)
        sheet.write(8, 14, 'Plan',title_style1_table_head)
        sheet.write(8, 15, 'Premium',title_style1_table_head)
        sheet.write(8, 16, 'Premium Frequency',title_style1_table_head)
        sheet.write(8, 17, 'Capitation/FFS',title_style1_table_head)
        sheet.write(8, 18, 'Coverage',title_style1_table_head)
        sheet.write(8, 19, 'Plan Type',title_style1_table_head)
        sheet.write(8, 20, 'Adult Dependent Count',title_style1_table_head)
        sheet.write(8, 21, 'Child Dependent Count',title_style1_table_head)
        sheet.write(8, 22, 'Provider Name',title_style1_table_head)
        sheet.write(8, 23, 'Provider ID',title_style1_table_head)
        sheet.write(8, 24, 'Provider Location',title_style1_table_head)
        sheet.write(8, 25, 'Enrollee Location',title_style1_table_head)
        sheet.write(8, 26, 'Policy End Date',title_style1_table_head)
        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['surname'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['code'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['employer'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['cicode'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['policy'], title_style1_table_data_sub)
            sheet.write(row_data, 7, line['dob'], title_style1_table_data_sub)
            sheet.write(row_data, 8, line['gender'], title_style1_table_data_sub)
            sheet.write(row_data, 9, line['marital'], title_style1_table_data_sub)
            sheet.write(row_data, 10, line['start_date'], title_style1_table_data_sub)
            sheet.write(row_data, 11, line['mthyr'], title_style1_table_data_sub)
            sheet.write(row_data, 12, line['end_date'], title_style1_table_data_sub)
            sheet.write(row_data, 13, line['active'], title_style1_table_data_sub)
            sheet.write(row_data, 14, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 15, line['price_unit'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 16, line['duration'], title_style1_table_data)
            sheet.write(row_data, 17, line['cap_ffs'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 18, line['coverage'], title_style1_table_data_sub)
            sheet.write(row_data, 19, line['type'], title_style1_table_data_sub)
            sheet.write(row_data, 20, line['adult_dep_cnt'], title_style1_table_data)
            sheet.write(row_data, 21, line['child_deps_cnt'], title_style1_table_data)
            sheet.write(row_data, 22, line['hcp'], title_style1_table_data)
            sheet.write(row_data, 23, str(line['hcpid']), title_style1_table_data)
            sheet.write(row_data, 24, line['town'], title_style1_table_data_sub)
            sheet.write(row_data, 25, line['city'], title_style1_table_data_sub)
            sheet.write(row_data, 26, line['pend_date'], title_style1_table_data_sub)
            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Members.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel2(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Claims Paid'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)       
        
        sheet.write(8, 0, 'Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Treatment Date',title_style1_table_head)
        sheet.write(8, 5, 'Date of Birth',title_style1_table_head)
        sheet.write(8, 6, 'Gender',title_style1_table_head)
        sheet.write(8, 7, 'Occupation',title_style1_table_head)
        sheet.write(8, 8, 'Plan Type',title_style1_table_head)
        sheet.write(8, 9, 'Business Category',title_style1_table_head)
        sheet.write(8, 10, 'Health Provider',title_style1_table_head)
        sheet.write(8, 11, 'Employer/Sector',title_style1_table_head)
        sheet.write(8, 12, 'Diagnosis',title_style1_table_head)
        sheet.write(8, 13, 'Diagnosis Code',title_style1_table_head)
        sheet.write(8, 14, 'In/Out Hospital',title_style1_table_head)
        sheet.write(8, 15, 'No of Days of Admission',title_style1_table_head)
        sheet.write(8, 16, 'Treatment/Referral',title_style1_table_head)
        sheet.write(8, 17, 'Policy Number',title_style1_table_head)
        sheet.write(8, 18, 'Chronic/Non-Chronic',title_style1_table_head)
        sheet.write(8, 19, 'Cover Period(Start)',title_style1_table_head)
        sheet.write(8, 20, 'Cover Period(End)',title_style1_table_head)
        sheet.write(8, 21, 'Claims Incurred Date',title_style1_table_head)
        sheet.write(8, 22, 'Claims Notification Date',title_style1_table_head)
        sheet.write(8, 23, 'Date of Approval',title_style1_table_head)
        sheet.write(8, 24, 'Claims Paid Date',title_style1_table_head)
        sheet.write(8, 25, 'Provider Bill',title_style1_table_head)
        sheet.write(8, 26, 'Approved Claims Amount',title_style1_table_head)
        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['code'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['dob'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['gender'], title_style1_table_data_sub)
            sheet.write(row_data, 7, line['occupation'], title_style1_table_data_sub)
            sheet.write(row_data, 8, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 9, line['categ'], title_style1_table_data_sub)
            sheet.write(row_data, 10, line['hcp'], title_style1_table_data)
            sheet.write(row_data, 11, line['employer'], title_style1_table_data)
            sheet.write(row_data, 12, line['diagnosis'], title_style1_table_data)
            sheet.write(row_data, 13, line['ccode'], title_style1_table_data)
            sheet.write(row_data, 14, line['inoutw'], title_style1_table_data)
            sheet.write(row_data, 15, line['duration'], title_style1_table_data)
            sheet.write(row_data, 16, line['referal'], title_style1_table_data)
            sheet.write(row_data, 17, line['policy'], title_style1_table_data_sub)
            sheet.write(row_data, 18, line['chronic'], title_style1_table_data_sub)
            sheet.write(row_data, 19, line['start_date'], title_style1_table_data_sub)
            sheet.write(row_data, 20, line['end_date'], title_style1_table_data_sub)
            sheet.write(row_data, 21, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 22, line['receipt_date'], title_style1_table_data_sub)
            sheet.write(row_data, 23, line['write_date'], title_style1_table_data_sub)
            sheet.write(row_data, 24, line['payment_date'], title_style1_table_data_sub)
            sheet.write(row_data, 25, line['provider_total'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 26, line['amount_total'], title_style1_table_data_sub_amount)

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Paid.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }
		

    def _print_report_excel3(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Claims Payment'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
        
       
        
        sheet.write(8, 0, 'Enrollee Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Claims Code',title_style1_table_head)
        sheet.write(8, 5, 'Treatment Date',title_style1_table_head)
        sheet.write(8, 6, 'Date of Receipt of Bill',title_style1_table_head)
        sheet.write(8, 7, 'Provider ID',title_style1_table_head)
        sheet.write(8, 8, 'Drugs/Service',title_style1_table_head)
        sheet.write(8, 9, 'Employer',title_style1_table_head)
        sheet.write(8, 10, 'Amount',title_style1_table_head)
        sheet.write(8, 11, 'Invoice Status',title_style1_table_head)
        sheet.write(8, 12, 'Date of Payment',title_style1_table_head)

        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['enrollee'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['claim'], title_style1_table_data)
            sheet.write(row_data, 5, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['receipt_date'], title_style1_table_data_sub)
            sheet.write(row_data, 7, line['hcp'], title_style1_table_data)
            sheet.write(row_data, 8, line['drug'], title_style1_table_data)
            sheet.write(row_data, 9, line['employer'], title_style1_table_data)
            sheet.write(row_data, 10, line['price_subtotal'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 11, line['state'], title_style1_table_data)
            sheet.write(row_data, 12, line['payment_date'], title_style1_table_data_sub)


            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Payments.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel4(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Claims Details_Procedures'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
        
        sheet.write(8, 0, 'Enrollee Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Claims Code',title_style1_table_head)
        sheet.write(8, 5, 'Treatment Date',title_style1_table_head)
        sheet.write(8, 6, 'Drugs/Service',title_style1_table_head)
        sheet.write(8, 7, 'Employer',title_style1_table_head)
        sheet.write(8, 8, 'Amount',title_style1_table_head)

        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['enrollee'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['claim'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['drug'], title_style1_table_data_sub)
            sheet.write(row_data, 7, line['employer'], title_style1_table_data)
            sheet.write(row_data, 8, line['price_total'], title_style1_table_data_sub_amount)

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Details_Procedures.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }


    def _print_report_excel5(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Diagnosis'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
        
        sheet.write(8, 0, 'Enrollee Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Claims Code',title_style1_table_head)
        sheet.write(8, 5, 'Date',title_style1_table_head)
        sheet.write(8, 6, 'Type',title_style1_table_head)
        sheet.write(8, 7, 'Diagnosis',title_style1_table_head)
        sheet.write(8, 8, 'Details',title_style1_table_head)

        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['enrollee'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['claim'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 6, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 7, line['diagnosis'], title_style1_table_data_sub)
            sheet.write(row_data, 8, line['details'], title_style1_table_data_sub)

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Diagnosis.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel_cap(self, datas):
        workbook = Workbook()        
   
        sheet = workbook.active
        sheet.title="Premium Data"
        
        sheet.merge_cells('A1:G2')
        cell = sheet.cell(row=1,column=1)
        cell.value = 'Actuary Report'
        #cell.style = title
        
        sheet['I1'] =  'Printing Date: ' + datetime.now().strftime('%Y-%m-%d')
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.cell(row=4, column=1).value = 'Company'
        sheet.cell(row=5, column=1).value = comp_id.name
		
        sheet.cell(row=9, column=1).value = 'Enrollee Code'
        sheet.cell(row=9, column=2).value = 'Surname'
        sheet.cell(row=9, column=3).value = 'Firstname'
        sheet.cell(row=9, column=4).value = 'Othername'
        sheet.cell(row=9, column=5).value = 'Name of Company'
        sheet.cell(row=9, column=6).value = 'Enrollee Policy Number'
        sheet.cell(row=9, column=7).value = 'Company ID'
        sheet.cell(row=9, column=8).value = 'Plan Type'
        sheet.cell(row=9, column=9).value = 'Provider Name'
        sheet.cell(row=9, column=10).value = 'Provider ID'
        sheet.cell(row=9, column=11).value = 'Capitation Amount'

        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.cell(row=row_data, column=1).value = line['code']
            sheet.cell(row=row_data, column=2).value = line['surname']
            sheet.cell(row=row_data, column=3).value = line['firstname']
            sheet.cell(row=row_data, column=4).value = line['othername']
            sheet.cell(row=row_data, column=5).value = line['employer']
            sheet.cell(row=row_data, column=6).value = line['policy']
            sheet.cell(row=row_data, column=7).value = line['id']
            sheet.cell(row=row_data, column=8).value = self.replacer(str(line['name']))
            sheet.cell(row=row_data, column=9).value = line['hcp']
            sheet.cell(row=row_data, column=10).value = line['hid']
            sheet.cell(row=row_data, column=11).value = line['cap_amount']

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Capitation_Data.xlsx', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel_claim_out(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Claims Outstanding'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
              
        sheet.write(8, 0, 'Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Date of Birth',title_style1_table_head)
        sheet.write(8, 5, 'Gender',title_style1_table_head)
        sheet.write(8, 6, 'Occupation',title_style1_table_head)
        sheet.write(8, 7, 'Plan Type',title_style1_table_head)
        sheet.write(8, 8, 'Business Category',title_style1_table_head)
        sheet.write(8, 9, 'Health Provider',title_style1_table_head)
        sheet.write(8, 10, 'Employer/Sector',title_style1_table_head)
        sheet.write(8, 11, 'Diagnosis',title_style1_table_head)
        sheet.write(8, 12, 'Treatment/Referral',title_style1_table_head)
        sheet.write(8, 13, 'Policy Number',title_style1_table_head)
        sheet.write(8, 14, 'Cover Periods(Start)',title_style1_table_head)
        sheet.write(8, 15, 'Cover Periods(End)',title_style1_table_head)
        sheet.write(8, 16, 'Claims Incurred Date',title_style1_table_head)
        sheet.write(8, 17, 'Claims Notification Date',title_style1_table_head)
        sheet.write(8, 18, 'Total Outstanding Claims Amount',title_style1_table_head)
        sheet.write(8, 19, 'Diagnosis Code',title_style1_table_head)
        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['code'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['dob'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['gender'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['occupation'], title_style1_table_data_sub)
            sheet.write(row_data, 7, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 8, line['categ'], title_style1_table_data)
            sheet.write(row_data, 9, line['provider'], title_style1_table_data)
            sheet.write(row_data, 10, line['employer'], title_style1_table_data)
            sheet.write(row_data, 11, line['diagnosis'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 12, line['referal'], title_style1_table_data_sub)
            sheet.write(row_data, 13, line['policy'], title_style1_table_data_sub)
            sheet.write(row_data, 14, line['start_date'], title_style1_table_data_sub)
            sheet.write(row_data, 15, line['end_date'], title_style1_table_data_sub)
            sheet.write(row_data, 16, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 17, line['receipt_date'], title_style1_table_data_sub)
            sheet.write(row_data, 18, line['amount_total'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 19, line['ccode'], title_style1_table_data)

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Outstanding.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel_prem(self, datas):
        workbook = Workbook()
        #header = NamedStyle(name="header")
        #header.font = Font(bold=True)
        #header.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        #number_format = 'Comma [0]'
        #date_format = 'dd/MM/yyyy'
        #title = Font(name='Calibri',size=20,bold=True)
        
        #sheet_name = 'Premium Data'
        sheet = workbook.active
        sheet.title="Premium Data"
        
        sheet.merge_cells('A1:G2')
        cell = sheet.cell(row=1,column=1)
        cell.value = 'Actuary Report'
        #cell.style = title
        
        sheet['I1'] =  'Printing Date: ' + datetime.now().strftime('%Y-%m-%d')
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.cell(row=4, column=1).value = 'Company'
        sheet.cell(row=5, column=1).value = comp_id.name
              
        sheet.cell(row=9, column=1).value = 'Code'
        sheet.cell(row=9, column=2).value = 'Surname'
        sheet.cell(row=9, column=3).value = 'Firstname'
        sheet.cell(row=9, column=4).value = 'Othername'
        sheet.cell(row=9, column=5).value = 'Date of Birth'
        sheet.cell(row=9, column=6).value = 'Gender'
        sheet.cell(row=9, column=7).value = 'Occupation'
        sheet.cell(row=9, column=8).value = 'Plan Type'
        sheet.cell(row=9, column=9).value = 'Business Category'
        sheet.cell(row=9, column=10).value = 'Health Provider'
        sheet.cell(row=9, column=11).value = 'Employer/Sector'
        sheet.cell(row=9, column=12).value = 'Policy Number'
        sheet.cell(row=9, column=13).value = 'Cover Period(Start)'
        sheet.cell(row=9, column=14).value = 'Cover Period(End)'
        sheet.cell(row=9, column=15).value = 'Gross Premium Amount'

        roww = 10       
        row_data = roww+1
        for line in datas:
            sheet.cell(row=row_data, column=1).value = line['code']
            sheet.cell(row=row_data, column=2).value = line['surname']
            sheet.cell(row=row_data, column=3).value = line['firstname']
            sheet.cell(row=row_data, column=4).value = line['othername']
            sheet.cell(row=row_data, column=5).value = line['dob']
            sheet.cell(row=row_data, column=6).value = line['gender']
            sheet.cell(row=row_data, column=7).value = line['occupation']
            sheet.cell(row=row_data, column=8).value = line['type']
            sheet.cell(row=row_data, column=9).value = line['categ']
            sheet.cell(row=row_data, column=10).value = line['provider']
            sheet.cell(row=row_data, column=11).value = line['employer']
            sheet.cell(row=row_data, column=12).value = line['policy']
            sheet.cell(row=row_data, column=13).value = line['start_date']
            sheet.cell(row=row_data, column=14).value = line['end_date']
            sheet.cell(row=row_data, column=15).value = line['price_subtotal']

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Premium.xlsx', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            #'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel_claim_incurred(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        sheet_name = 'Incurred Claims/Refunds'
        sheet = workbook.add_sheet(sheet_name)
        
        sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        comp_id = self.env.user.company_id
        currency_id = comp_id.currency_id
        sheet.write(3, 0, 'Company',title_style1_table_head1)
        sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
              
        sheet.write(8, 0, 'Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Date of Birth',title_style1_table_head)
        sheet.write(8, 5, 'Gender',title_style1_table_head)
        sheet.write(8, 6, 'Occupation',title_style1_table_head)
        sheet.write(8, 7, 'Plan Type',title_style1_table_head)
        sheet.write(8, 8, 'Business Category',title_style1_table_head)
        sheet.write(8, 9, 'Health Provider',title_style1_table_head)
        sheet.write(8, 10, 'Employer/Sector',title_style1_table_head)
        sheet.write(8, 11, 'Diagnosis',title_style1_table_head)
        sheet.write(8, 12, 'Treatment/Referral',title_style1_table_head)
        sheet.write(8, 13, 'Policy Number',title_style1_table_head)
        sheet.write(8, 14, 'Cover Periods(Start)',title_style1_table_head)
        sheet.write(8, 15, 'Cover Periods(End)',title_style1_table_head)
        sheet.write(8, 16, 'Claims Incurred Date',title_style1_table_head)
        sheet.write(8, 17, 'Claims Notification Date',title_style1_table_head)
        sheet.write(8, 18, 'Total Outstanding Claims Amount',title_style1_table_head)
        sheet.write(8, 19, 'Diagnosis Code',title_style1_table_head)
        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['code'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['dob'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['gender'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['occupation'], title_style1_table_data_sub)
            sheet.write(row_data, 7, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 8, line['categ'], title_style1_table_data)
            sheet.write(row_data, 9, line['provider'], title_style1_table_data)
            sheet.write(row_data, 10, line['employer'], title_style1_table_data)
            sheet.write(row_data, 11, line['diagnosis'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 12, line['referal'], title_style1_table_data_sub)
            sheet.write(row_data, 13, line['policy'], title_style1_table_data_sub)
            sheet.write(row_data, 14, line['start_date'], title_style1_table_data_sub)
            sheet.write(row_data, 15, line['end_date'], title_style1_table_data_sub)
            sheet.write(row_data, 16, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 17, line['receipt_date'], title_style1_table_data_sub)
            sheet.write(row_data, 18, line['amount_total'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 19, line['ccode'], title_style1_table_data_sub)

            row_data = row_data + 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Incurred.xls', 'xls_output': base64.encodestring(stream.getvalue())})
        return {
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }
        
    def _print_report_excel_claim_incurred_xlsx(self, datas):
        output = io.BytesIO()
        #workbook = xlsxwriter.Workbook(output)
        workbook = xlsxwriter.Workbook(output,{'in_memory': True})
        sheet = workbook.add_worksheet('Actuary Report')
        title_style1_table_head = workbook.add_format({'bold': True})
        number = workbook.add_format({'num_format': '#,##0'})
        header = workbook.add_format({'bold': True, 'font_size': 16})
        
                
        #sheet_name = 'Incurred Claims/Refunds'
        #sheet = workbook.add_sheet(sheet_name)
        
        #sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
        #sheet.write(0, 8, 'Printing Date: '+datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
        
        #comp_id = self.env.user.company_id
        #currency_id = comp_id.currency_id
        #sheet.write(3, 0, 'Company',title_style1_table_head1)
        #sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
        sheet.set_column(1, 20, 25)
        sheet.write(8, 0, 'Code',title_style1_table_head)
        sheet.write(8, 1, 'Surname',title_style1_table_head)
        sheet.write(8, 2, 'Firstname',title_style1_table_head)
        sheet.write(8, 3, 'Othername',title_style1_table_head)
        sheet.write(8, 4, 'Date of Birth',title_style1_table_head)
        sheet.write(8, 5, 'Gender',title_style1_table_head)
        sheet.write(8, 6, 'Occupation',title_style1_table_head)
        sheet.write(8, 7, 'Plan Type',title_style1_table_head)
        sheet.write(8, 8, 'Business Category',title_style1_table_head)
        sheet.write(8, 9, 'Health Provider',title_style1_table_head)
        sheet.write(8, 10, 'Employer/Sector',title_style1_table_head)
        sheet.write(8, 11, 'Diagnosis',title_style1_table_head)
        sheet.write(8, 12, 'Treatment/Referral',title_style1_table_head)
        sheet.write(8, 13, 'Policy Number',title_style1_table_head)
        sheet.write(8, 14, 'Cover Periods(Start)',title_style1_table_head)
        sheet.write(8, 15, 'Cover Periods(End)',title_style1_table_head)
        sheet.write(8, 16, 'Claims Incurred Date',title_style1_table_head)
        sheet.write(8, 17, 'Claims Notification Date',title_style1_table_head)
        sheet.write(8, 18, 'Total Outstanding Claims Amount',title_style1_table_head)
        sheet.write(8, 19, 'Diagnosis Code',title_style1_table_head)
        roww = 9        
        row_data = roww+1
        for line in datas:
            sheet.write(row_data, 0, line['code'])
            sheet.write(row_data, 1, line['surname'])
            sheet.write(row_data, 2, line['firstname'])
            sheet.write(row_data, 3, line['othername'])
            sheet.write(row_data, 4, line['dob'])
            sheet.write(row_data, 5, line['gender'])
            sheet.write(row_data, 6, line['occupation'])
            sheet.write(row_data, 7, self.replacer(str(line['name'])))
            sheet.write(row_data, 8, line['categ'])
            sheet.write(row_data, 9, line['provider'])
            sheet.write(row_data, 10, line['employer'])
            sheet.write(row_data, 11, line['diagnosis'])
            sheet.write(row_data, 12, line['referal'])
            sheet.write(row_data, 13, line['policy'])
            sheet.write(row_data, 14, line['start_date'])
            sheet.write(row_data, 15, line['end_date'])
            sheet.write(row_data, 16, line['date_order'])
            sheet.write(row_data, 17, line['receipt_date'])
            sheet.write(row_data, 18, line['amount_total'],number)
            sheet.write(row_data, 19, line['ccode'])

            row_data = row_data + 1
        roww = row_data + 3
        workbook.close
        output.seek(0)
        #response.stream.write(output.read())
        #output.close()
        #with open('Incurred_Claims', "rb") as file:
        #    file_base64 = base64.b64encode(file.read())
        
      
        data = base64.encodestring(output.getvalue())
        output.close()
        raise UserError(_('%s') % roww)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Incurred.xlsx', 'xls_output': data})
        return {
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }

    def _print_report_excel_claim_incurred_multi(self, datas):
        workbook = xlwt.Workbook()
        title_style_comp = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold off, italic off, height 450')
        title_style_comp_left = xlwt.easyxf('align: horiz left ; font: name Times New Roman,bold off, italic off, height 450')
        title_style = xlwt.easyxf('align: horiz center ;font: name Times New Roman,bold off, italic off, height 350')
        title_style2 = xlwt.easyxf('font: name Times New Roman, height 200')
        title_style1 = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190; borders: top double, bottom double, left double, right double;')
        title_style1_table_head = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head1 = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200')
        title_style1_consultant = xlwt.easyxf('font: name Times New Roman,bold on, italic off, height 200; borders: top double, bottom double, left double, right double;')
        title_style1_table_head_center = xlwt.easyxf('align: horiz center ; font: name Times New Roman,bold on, italic off, height 190; borders: top thick, bottom thick, left thick, right thick;')

        title_style1_table_data = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub = xlwt.easyxf('font: name Times New Roman,bold off, italic off, height 190',num_format_str='dd/MM/yyyy')
        #title_style1_table_data_sub.num_format_str = '#,##.00'
        title_style1_table_data_sub_amount = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        title_style1_table_data_sub_balance = xlwt.easyxf('align: horiz right ;font: name Times New Roman,bold off, italic off, height 190',num_format_str='#,##0.00')
        #gl_report_obj = self.env['report.account_extra_reports.report_partnerledger']
        
        row_count = 1
        page = 0
                    
        for line in datas:
            if row_count == 1:
                page+=1
                row_count = 4
                sheet_name = 'Incurred Claims/Refunds' + ' ' + str(page)
                sheet = workbook.add_sheet(sheet_name)
                
                sheet.write_merge(0, 1, 0, 6, 'Actuary Report', title_style_comp_left)
                sheet.write(0, 8, 'Printing Date: ' + datetime.now().strftime('%Y-%m-%d'), title_style1_table_head1)
                sheet.write(0,12, 'Page ' + str(page))
                
                comp_id = self.env.user.company_id
                currency_id = comp_id.currency_id
                sheet.write(3, 0, 'Company',title_style1_table_head1)
                sheet.write(4, 0, comp_id.name, title_style1_table_data_sub)
                      
                sheet.write(8, 0, 'Code',title_style1_table_head)
                sheet.write(8, 1, 'Surname',title_style1_table_head)
                sheet.write(8, 2, 'Firstname',title_style1_table_head)
                sheet.write(8, 3, 'Othername',title_style1_table_head)
                sheet.write(8, 4, 'Date of Birth',title_style1_table_head)
                sheet.write(8, 5, 'Gender',title_style1_table_head)
                sheet.write(8, 6, 'Occupation',title_style1_table_head)
                sheet.write(8, 7, 'Plan Type',title_style1_table_head)
                sheet.write(8, 8, 'Business Category',title_style1_table_head)
                sheet.write(8, 9, 'Health Provider',title_style1_table_head)
                sheet.write(8, 10, 'Employer/Sector',title_style1_table_head)
                sheet.write(8, 11, 'Diagnosis',title_style1_table_head)
                sheet.write(8, 12, 'Treatment/Referral',title_style1_table_head)
                sheet.write(8, 13, 'Policy Number',title_style1_table_head)
                sheet.write(8, 14, 'Cover Periods(Start)',title_style1_table_head)
                sheet.write(8, 15, 'Cover Periods(End)',title_style1_table_head)
                sheet.write(8, 16, 'Claims Incurred Date',title_style1_table_head)
                sheet.write(8, 17, 'Claims Notification Date',title_style1_table_head)
                sheet.write(8, 18, 'Total Outstanding Claims Amount',title_style1_table_head)
                sheet.write(8, 19, 'Diagnosis Code',title_style1_table_head)
                roww = 9        
                row_data = roww+1
                
            sheet.write(row_data, 0, line['code'], title_style1_table_data_sub)
            column = sheet.col(0)
            column.width = 256 * 25
            sheet.write(row_data, 1, line['surname'], title_style1_table_data_sub)
            sheet.write(row_data, 2, line['firstname'], title_style1_table_data_sub)
            sheet.write(row_data, 3, line['othername'], title_style1_table_data_sub)
            sheet.write(row_data, 4, line['dob'], title_style1_table_data_sub)
            sheet.write(row_data, 5, line['gender'], title_style1_table_data_sub)
            sheet.write(row_data, 6, line['occupation'], title_style1_table_data_sub)
            sheet.write(row_data, 7, self.replacer(str(line['name'])), title_style1_table_data_sub)
            sheet.write(row_data, 8, line['categ'], title_style1_table_data)
            sheet.write(row_data, 9, line['provider'], title_style1_table_data)
            sheet.write(row_data, 10, line['employer'], title_style1_table_data)
            sheet.write(row_data, 11, line['diagnosis'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 12, line['referal'], title_style1_table_data_sub)
            sheet.write(row_data, 13, line['policy'], title_style1_table_data_sub)
            sheet.write(row_data, 14, line['start_date'], title_style1_table_data_sub)
            sheet.write(row_data, 15, line['end_date'], title_style1_table_data_sub)
            sheet.write(row_data, 16, line['date_order'], title_style1_table_data_sub)
            sheet.write(row_data, 17, line['receipt_date'], title_style1_table_data_sub)
            sheet.write(row_data, 18, line['amount_total'], title_style1_table_data_sub_amount)
            sheet.write(row_data, 19, line['ccode'], title_style1_table_data_sub)

            row_data = row_data + 1
            if row_data == 65001:
                row_count = 1
        roww = row_data + 3

        stream = io.BytesIO()
        workbook.save(stream)
        attach_id = self.env['actuary.report.output.wizard'].create({'name':'Actuary_Claims_Incurred.xls', 'xls_output': base64.encodebytes(stream.getvalue())})
        return {
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'actuary.report.output.wizard',
            'res_id':attach_id.id,
            'type': 'ir.actions.act_window',
            'target':'new'
        }
        
class ActuaryReportOutputWizard(models.Model):
    _name = 'actuary.report.output.wizard'
    _description = 'Wizard to store the Excel output'

    xls_output = fields.Binary(string='Excel Output',readonly=True)
    name = fields.Char(
        string='File Name',
        help='Save report as .xls format',
        default='Actuary.xlsx',
    )
